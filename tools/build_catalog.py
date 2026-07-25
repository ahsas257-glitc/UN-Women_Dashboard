from __future__ import annotations

import json
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
XLSFORM_DIR = ROOT / "XLS_Forms"
OUTPUT = ROOT / "data" / "metadata" / "xlsform_catalog.json"


def clean(value):
    return value.strip() if isinstance(value, str) else value


def sheet_records(
    workbook: openpyxl.Workbook,
    sheet_name: str,
    allowed: set[str],
) -> list[dict]:
    if sheet_name not in workbook.sheetnames:
        return []
    rows = list(workbook[sheet_name].iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    return [
        {
            header: clean(values[index]) if index < len(values) else None
            for index, header in enumerate(headers)
            if header in allowed
        }
        for values in rows[1:]
    ]


def build_form(path: Path) -> dict:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    question_fields = {
        "type",
        "name",
        "label::English (en)",
        "label::English",
        "label",
        "required",
        "relevant",
        "constraint",
        "constraint_message::English (en)",
        "calculation",
    }
    choice_fields = {
        "list_name",
        "name",
        "label::English (en)",
        "label::English",
        "label",
    }
    questions = [
        row
        for row in sheet_records(workbook, "survey", question_fields)
        if row.get("type") and row.get("name")
    ]
    choices = [
        row
        for row in sheet_records(workbook, "choices", choice_fields)
        if row.get("list_name") and row.get("name") is not None
    ]
    workbook.close()
    return {
        "file": str(Path("XLS_Forms") / path.name),
        "questions": questions,
        "choices": choices,
    }


def main() -> None:
    forms = [
        build_form(path)
        for path in sorted(XLSFORM_DIR.glob("*.xlsx"))
        if not path.name.startswith("~$")
    ]
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(
        json.dumps(forms, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    print(f"Wrote {len(forms)} questionnaire definitions to {OUTPUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
