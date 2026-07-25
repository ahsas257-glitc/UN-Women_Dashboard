from __future__ import annotations

import json
import re
from collections import defaultdict
from functools import lru_cache
from pathlib import Path
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    import openpyxl

from src.config import CATALOG_PATH, FORM_TITLES, XLSFORM_DIR


NON_ANALYTIC_TYPES = {
    "text",
    "note",
    "begin_group",
    "end_group",
    "begin_repeat",
    "end_repeat",
    "start",
    "end",
    "today",
    "deviceid",
    "username",
    "start-geopoint",
    "geopoint",
    "image",
    "audio",
    "video",
}

TECHNICAL_FIELD_NAMES = {
    "submission_uuid",
    "instance_name",
    "instancename",
    "form_id",
    "form_code",
    "form_version",
    "form_title",
    "form_short_name",
}


def normalize_label(value: Any) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"\s+", " ", text).strip().casefold()


def form_code_from_filename(path: str) -> str | None:
    # Metadata may be generated on Windows but consumed on Linux in Streamlit
    # Cloud. Split both path separators instead of delegating to the host OS.
    filename = re.split(r"[\\/]", str(path))[-1]
    match = re.match(r"(C[1-7]|F0[1-6])", filename, flags=re.I)
    return match.group(1).upper() if match else None


def question_label(question: dict[str, Any]) -> str:
    return (
        question.get("label::English (en)")
        or question.get("label::English")
        or question.get("label")
        or question.get("name")
        or "Unnamed field"
    )


def base_type(question_type: Any) -> str:
    return str(question_type or "").strip().split(" ", 1)[0]


def _clean_cell(value: Any) -> Any:
    return value.strip() if isinstance(value, str) else value


def _read_sheet_records(workbook: openpyxl.Workbook, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    rows = list(workbook[sheet_name].iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    return [
        {
            header: _clean_cell(values[index]) if index < len(values) else None
            for index, header in enumerate(headers)
            if header
        }
        for values in rows[1:]
    ]


def _load_xlsforms(directory: Path) -> list[dict[str, Any]]:
    import openpyxl

    forms = []
    for path in sorted(directory.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        questions = [
            row
            for row in _read_sheet_records(workbook, "survey")
            if row.get("type") and row.get("name")
        ]
        choices = [
            row
            for row in _read_sheet_records(workbook, "choices")
            if row.get("list_name") and row.get("name") is not None
        ]
        workbook.close()
        forms.append({"file": str(path), "questions": questions, "choices": choices})
    return forms


@lru_cache(maxsize=1)
def load_catalog(
    path: str = str(CATALOG_PATH),
    xlsform_dir: str = str(XLSFORM_DIR),
) -> dict[str, dict[str, Any]]:
    directory = Path(xlsform_dir)
    metadata_path = Path(path)
    if metadata_path.is_file():
        raw = json.loads(Path(path).read_text(encoding="utf-8"))
    elif directory.is_dir() and any(directory.glob("*.xlsx")):
        raw = _load_xlsforms(directory)
    else:
        raise FileNotFoundError(
            "Questionnaire metadata is unavailable. "
            "Include data/metadata/xlsform_catalog.json or XLS_Forms/*.xlsx."
        )
    catalog: dict[str, dict[str, Any]] = {}
    for form in raw:
        source_file = str(form["file"]).replace("\\", "/")
        code = form_code_from_filename(source_file)
        if not code:
            continue
        choices_by_list: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for choice in form.get("choices", []):
            choices_by_list[str(choice.get("list_name"))].append(choice)

        questions = []
        by_name = {}
        by_label = {}
        for index, question in enumerate(form.get("questions", [])):
            item = dict(question)
            item["_index"] = index
            item["_base_type"] = base_type(item.get("type"))
            item["_display_label"] = question_label(item)
            questions.append(item)
            if item.get("name"):
                by_name[normalize_label(item["name"])] = item
            by_label[normalize_label(item["_display_label"])] = item

        catalog[code] = {
            "code": code,
            "title": FORM_TITLES.get(code, code),
            "source_file": source_file,
            "questions": questions,
            "choices_by_list": dict(choices_by_list),
            "by_name": by_name,
            "by_label": by_label,
        }
    return catalog


def resolve_question(form_code: str, column: str) -> dict[str, Any] | None:
    form = load_catalog().get(form_code, {})
    key = normalize_label(column)
    return form.get("by_label", {}).get(key) or form.get("by_name", {}).get(key)


def is_analytic_question(question: dict[str, Any] | None) -> bool:
    if not question:
        return False
    name = normalize_label(question.get("name")).replace(" ", "_")
    return (
        question.get("_base_type") not in NON_ANALYTIC_TYPES
        and name not in TECHNICAL_FIELD_NAMES
        and not name.startswith("_")
    )


def analytic_questions(form_code: str) -> list[dict[str, Any]]:
    form = load_catalog().get(form_code, {})
    return [q for q in form.get("questions", []) if is_analytic_question(q)]


def score_questions(form_code: str) -> list[dict[str, Any]]:
    form = load_catalog().get(form_code, {})
    questions = form.get("questions", [])
    result = []
    for question in questions:
        name = str(question.get("name") or "").lower()
        if "score" not in name:
            continue
        if any(
            token in name
            for token in (
                "guide",
                "target",
                "avg",
                "average",
                "mean",
                "gap",
                "change",
                "percent",
                "count",
                "sum",
                "total",
            )
        ):
            continue
        if question.get("_base_type") not in {"calculate", "select_one", "integer", "decimal"}:
            continue
        display = question["_display_label"]
        if question.get("_base_type") == "calculate" and display == question.get("name"):
            for prior in reversed(questions[: question["_index"]]):
                if (
                    prior.get("_base_type") in {"select_one", "select_multiple", "integer", "decimal"}
                    and prior.get("_display_label")
                ):
                    display = prior["_display_label"]
                    break
        enriched = dict(question)
        enriched["_score_label"] = display
        result.append(enriched)
    return result
